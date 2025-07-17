import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import re
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from aplicacion.conectiondb.conexion import DBConnection

class ReporteDeCargas(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.db = DBConnection()
        self.create_widgets(parent)

    def create_widgets(self,parent):
        frame_reporte = ctk.CTkFrame(parent)
        frame_reporte.pack(pady=20, padx=20, fill="both", expand=True)

        label_titulo = ctk.CTkLabel(frame_reporte, text="Reporte de Cargas", font=("Arial", 24))
        label_titulo.pack(pady=10)

        self.treeview = ttk.Treeview(frame_reporte, columns=("NUMERO", "IDENTIFICACION", "FECHA"), show="headings", selectmode="extended")
        self.treeview.heading("NUMERO", text="Número")
        self.treeview.heading("IDENTIFICACION", text="Identificación")
        self.treeview.heading("FECHA", text="Fecha")
        self.treeview.pack(pady=10, fill="both", expand=True)

        button_ejecutar = ctk.CTkButton(frame_reporte, text="Ejecutar Consulta", command=self.ejecutar_consulta)
        button_ejecutar.pack(pady=10)

        self.cargar_datos_treeview()

    def procesar_vol(self, vol_list):
        if not vol_list:
            return 0
        resultado = []
        for vol in vol_list:
            numeros = re.findall(r'\d+', str(vol))
            resultado.append(int(numeros[0]) if numeros else 0)
        return sum(resultado) if len(resultado) > 1 else resultado[0]

    def apply_border_to_range(self, ws, start_row, end_row, start_col, end_col):
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                border = Border(
                    top=Side(style='thin') if row == start_row else None,
                    bottom=Side(style='thin') if row == end_row else None,
                    left=Side(style='thin') if col == start_col else None,
                    right=Side(style='thin') if col == end_col else None
                )
                cell.border = border

    def format_worksheet(self, ws, df):
        valor_col = df.columns.get_loc('VALOR') + 1
        for row in range(2, len(df) + 2):
            ws.cell(row=row, column=valor_col).number_format = '#,##0'

        current_city = None
        current_group_start = 2
        city_groups = []

        for idx, row in df.iterrows():
            if current_city != row['CIUDAD']:
                if current_city is not None:
                    city_groups.append((current_group_start, idx + 1))
                current_city = row['CIUDAD']
                current_group_start = idx + 2

        if current_group_start <= len(df) + 1:
            city_groups.append((current_group_start, len(df) + 1))

        for start_row, end_row in city_groups:
            self.apply_border_to_range(ws, start_row, end_row, 1, len(df.columns))

        total_row = len(df) + 2
        ws.cell(row=total_row, column=1, value='TOTAL')
        for col_name in ['VOL', 'VALOR', 'PESO']:
            col_letter = get_column_letter(df.columns.get_loc(col_name) + 1)
            ws.cell(row=total_row, column=df.columns.get_loc(col_name) + 1, 
                    value=f'=SUM({col_letter}2:{col_letter}{total_row - 1})')
            if col_name == 'VALOR':
                ws.cell(row=total_row, column=df.columns.get_loc('VALOR') + 1).number_format = '#,##0'

    def create_summary_sheet(self, writer, resultados_excel):
        summary_data = [{
            'IDENTIFICADOR': k,
            'VOL': df['VOL'].sum(),
            'VALOR': df['VALOR'].sum(),
            'PESO': df['PESO'].sum()
        } for k, df in resultados_excel.items()]

        summary_df = pd.DataFrame(summary_data)
        total_row = pd.DataFrame([{
            'IDENTIFICADOR': 'TOTAL',
            'VOL': summary_df['VOL'].sum(),
            'VALOR': summary_df['VALOR'].sum(),
            'PESO': summary_df['PESO'].sum()
        }])
        summary_df = pd.concat([summary_df, total_row], ignore_index=True)
        summary_df.to_excel(writer, sheet_name='RESUMEN', index=False)

        ws = writer.sheets['RESUMEN']
        for row in range(2, len(summary_df) + 2):
            ws.cell(row=row, column=summary_df.columns.get_loc('VALOR') + 1).number_format = '#,##0'
        self.apply_border_to_range(ws, 1, len(summary_df) + 1, 1, len(summary_df.columns))

    def guardar_archivo(self, resultados_excel):
        fecha_actual = datetime.now().strftime('%Y-%m-%d')
        ruta_archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"Romaneio {fecha_actual}.xlsx",
            title="Guardar archivo"
        )
        if ruta_archivo:
            with pd.ExcelWriter(ruta_archivo, engine='openpyxl') as writer:
                for hoja, df in resultados_excel.items():
                    df.to_excel(writer, sheet_name=hoja[:31], index=False)
                    self.format_worksheet(writer.sheets[hoja[:31]], df)
                self.create_summary_sheet(writer, resultados_excel)
            messagebox.showinfo("Éxito", f"Consulta ejecutada y resultados guardados en: {ruta_archivo}")

    def ejecutar_consulta(self):
        selected_items = self.treeview.selection()
        numeros = [self.treeview.item(item)['values'][0] for item in selected_items]
        if not numeros:
            messagebox.showwarning("Advertencia", "Seleccione al menos un número.")
            return

        resultados_excel = {}
        con = self.db.connect_mysql()
        cursor = con.cursor()

        for numero in numeros:
            query = f"""
                select
                    c.id,
                    c.identidad,
                    e.id,
                    e.nombre as EMPRESAS,
                    p.numero,
                    p.factura,
                    p.condicion,
                    e.ciudad,
                    e.localentrega as DIA_DE_ENTREGA,
                    p.observacion as VOL,
                    cast(sum(pp.cantidad_final * pp.precio_principal) as float) as valor,
                    cast(sum(pp.cantidad_final * p2.peso) as float) as peso
                from
                    cargamentos c
                left join carga_pedido cp on
                    cp.cargamento = c.id
                    and cp.activo = 'S'
                left join pedidos p on
                    p.numero = cp.pedido
                left join producto_pedido pp on
                    pp.pedido = p.numero
                left join productos p2 on
                    p2.id = pp.item
                left join empresas e on
                    e.id = p.empresa
                where
                    c.id = '{numero}'
                group by
                    c.id,
                    c.identidad,
                    e.id,
                    e.nombre,
                    p.numero,
                    p.factura,
                    p.condicion,
                    e.ciudad,
                    e.localentrega,
                    p.observacion;
            """
            cursor.execute(query)
            resultados = cursor.fetchall()

            df_resultados = pd.DataFrame(resultados, columns=[
                'ROM', 'IDENTIFICACAO', 'CODIGO', 'EMPRESA',
                'PEDIDO', 'FACTURA', 'CONDICION',
                'CIUDAD', 'DIA_DE_ENTREGA',
                'VOL', 'VALOR', 'PESO']
            )
            df_resultados['VOL'] = df_resultados['VOL'].apply(self.procesar_vol)
            identificacion = df_resultados['IDENTIFICACAO'].iloc[0] if not df_resultados.empty else 'DESCONOCIDO'
            resultados_excel[f'{numero} {identificacion}'] = df_resultados

        self.guardar_archivo(resultados_excel)
        cursor.close()
        con.close()

    def cargar_datos_treeview(self):
        con = self.db.connect_mysql()
        cursor = con.cursor()
        cursor.execute("""
                       SELECT c.id, c.identidad,c.fecha_creacion FROM cargamentos c 
                        WHERE c.activo = 'S'
                        ORDER BY c.id DESC""")
        datos = cursor.fetchall()
        for numero, identificacion, fecha in datos:
            fecha_formateada = fecha.strftime("%d/%m/%Y") if fecha else ""
            self.treeview.insert("", "end", values=(numero, identificacion, fecha_formateada))
        cursor.close()
        con.close()
